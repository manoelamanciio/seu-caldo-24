# -*- coding: utf-8 -*-
"""
Módulo de Relatórios.
Sistema Seu Caldo 24 - M.A Sistemas.
"""

import datetime
from pathlib import Path

from PySide6.QtCore import QDate
from PySide6.QtWidgets import (
    QFileDialog,
    QDateEdit,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from config import EMPRESA
from services.relatorios_service import RelatoriosService
from utils.helpers import alerta, fmt_data, fmt_moeda, info


def _configurar_tabela(tabela):
    tabela.horizontalHeader().setSectionResizeMode(
        QHeaderView.Stretch
    )
    tabela.setEditTriggers(QTableWidget.NoEditTriggers)
    tabela.setSelectionBehavior(QTableWidget.SelectRows)


class RelatoriosWidget(QWidget):
    def __init__(self, usuario):
        super().__init__()

        self.usuario = usuario
        self.service = RelatoriosService()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)

        titulo = QLabel("Relatórios Gerenciais")
        titulo.setObjectName("tituloPagina")
        layout.addWidget(titulo)

        filtros = QHBoxLayout()

        self.inicio = QDateEdit()
        self.inicio.setCalendarPopup(True)
        self.inicio.setDate(
            QDate.currentDate().addMonths(-1)
        )

        self.fim = QDateEdit()
        self.fim.setCalendarPopup(True)
        self.fim.setDate(QDate.currentDate())

        atualizar = QPushButton("Atualizar")
        atualizar.clicked.connect(self.atualizar)

        exportar_excel = QPushButton("Exportar Excel")
        exportar_excel.setObjectName("btnSecundario")
        exportar_excel.clicked.connect(self._exportar_excel)

        exportar_pdf = QPushButton("Gerar PDF")
        exportar_pdf.setObjectName("btnSecundario")
        exportar_pdf.clicked.connect(self._exportar_pdf)

        filtros.addWidget(QLabel("Início:"))
        filtros.addWidget(self.inicio)
        filtros.addWidget(QLabel("Fim:"))
        filtros.addWidget(self.fim)
        filtros.addWidget(atualizar)
        filtros.addStretch()
        filtros.addWidget(exportar_excel)
        filtros.addWidget(exportar_pdf)

        layout.addLayout(filtros)

        self.tabs = QTabWidget()

        self.tabela_vendas = QTableWidget(0, 6)
        self.tabela_vendas.setHorizontalHeaderLabels([
            "Produto",
            "Quantidade",
            "Faturamento",
            "Custo",
            "Lucro",
            "Margem",
        ])
        _configurar_tabela(self.tabela_vendas)

        self.tabela_abc = QTableWidget(0, 6)
        self.tabela_abc.setHorizontalHeaderLabels([
            "Produto",
            "Faturamento",
            "% Total",
            "% Acumulado",
            "Classe",
            "Lucro",
        ])
        _configurar_tabela(self.tabela_abc)

        self.tabela_estoque = QTableWidget(0, 5)
        self.tabela_estoque.setHorizontalHeaderLabels([
            "Produto",
            "Categoria",
            "Setor",
            "Estoque atual",
            "Estoque mínimo",
        ])
        _configurar_tabela(self.tabela_estoque)

        self.tabela_financeiro = QTableWidget(0, 4)
        self.tabela_financeiro.setHorizontalHeaderLabels([
            "Indicador",
            "Valor",
            "Período inicial",
            "Período final",
        ])
        _configurar_tabela(self.tabela_financeiro)

        self.tabela_producao = QTableWidget(0, 7)
        self.tabela_producao.setHorizontalHeaderLabels([
            "Ordem",
            "Produto",
            "Planejado",
            "Produzido",
            "Perda",
            "Data",
            "Status",
        ])
        _configurar_tabela(self.tabela_producao)

        self.tabela_clientes = QTableWidget(0, 4)
        self.tabela_clientes.setHorizontalHeaderLabels([
            "Cliente",
            "Vendas",
            "Total comprado",
            "Ticket médio",
        ])
        _configurar_tabela(self.tabela_clientes)

        self.tabs.addTab(
            self.tabela_vendas,
            "Produtos / Rentabilidade"
        )
        self.tabs.addTab(
            self.tabela_abc,
            "Curva ABC"
        )
        self.tabs.addTab(
            self.tabela_estoque,
            "Estoque Baixo"
        )
        self.tabs.addTab(
            self.tabela_financeiro,
            "Financeiro"
        )
        self.tabs.addTab(
            self.tabela_producao,
            "Produção"
        )
        self.tabs.addTab(
            self.tabela_clientes,
            "Ranking de Clientes"
        )

        layout.addWidget(self.tabs)

        self.atualizar()

    def _periodo(self):
        inicio_data = self.inicio.date().toPython()
        fim_data = self.fim.date().toPython()

        inicio = datetime.datetime.combine(
            inicio_data,
            datetime.time.min,
        )
        fim = datetime.datetime.combine(
            fim_data,
            datetime.time.max,
        )

        return inicio, fim

    def atualizar(self):
        inicio, fim = self._periodo()

        self._atualizar_vendas(inicio, fim)
        self._atualizar_abc(inicio, fim)
        self._atualizar_estoque()
        self._atualizar_financeiro(inicio, fim)
        self._atualizar_producao(inicio, fim)
        self._atualizar_clientes(inicio, fim)

    def _atualizar_vendas(self, inicio, fim):
        dados = self.service.produtos_mais_vendidos(
            inicio,
            fim,
        )

        self.tabela_vendas.setRowCount(0)

        for item in dados:
            row = self.tabela_vendas.rowCount()
            self.tabela_vendas.insertRow(row)

            valores = [
                item["produto_nome"],
                item["quantidade"],
                fmt_moeda(item["faturamento"]),
                fmt_moeda(item["custo_total"]),
                fmt_moeda(item["lucro"]),
                f"{item['margem']:.2f}%",
            ]

            for coluna, valor in enumerate(valores):
                self.tabela_vendas.setItem(
                    row,
                    coluna,
                    QTableWidgetItem(str(valor)),
                )

    def _atualizar_abc(self, inicio, fim):
        dados = self.service.curva_abc(inicio, fim)

        self.tabela_abc.setRowCount(0)

        for item in dados:
            row = self.tabela_abc.rowCount()
            self.tabela_abc.insertRow(row)

            valores = [
                item["produto_nome"],
                fmt_moeda(item["faturamento"]),
                f"{item['percentual']:.2f}%",
                f"{item['percentual_acumulado']:.2f}%",
                item["classe"],
                fmt_moeda(item["lucro"]),
            ]

            for coluna, valor in enumerate(valores):
                self.tabela_abc.setItem(
                    row,
                    coluna,
                    QTableWidgetItem(str(valor)),
                )

    def _atualizar_estoque(self):
        dados = self.service.estoque_baixo()

        self.tabela_estoque.setRowCount(0)

        for produto in dados:
            row = self.tabela_estoque.rowCount()
            self.tabela_estoque.insertRow(row)

            valores = [
                produto.get("nome", ""),
                produto.get("categoria", ""),
                produto.get("setor", ""),
                produto.get("estoque_atual", 0),
                produto.get("estoque_minimo", 0),
            ]

            for coluna, valor in enumerate(valores):
                self.tabela_estoque.setItem(
                    row,
                    coluna,
                    QTableWidgetItem(str(valor)),
                )

    def _atualizar_financeiro(self, inicio, fim):
        resumo = self.service.resumo_financeiro(
            inicio,
            fim,
        )
        vendas = self.service.resumo_vendas(
            inicio,
            fim,
        )

        dados = [
            ("Entradas", resumo["entradas"]),
            ("Saídas", resumo["saidas"]),
            ("Saldo", resumo["saldo"]),
            ("Faturamento", vendas["faturamento"]),
            ("Quantidade de vendas", vendas["quantidade_vendas"]),
            ("Ticket médio", vendas["ticket_medio"]),
        ]

        self.tabela_financeiro.setRowCount(0)

        for indicador, valor in dados:
            row = self.tabela_financeiro.rowCount()
            self.tabela_financeiro.insertRow(row)

            valor_formatado = (
                str(valor)
                if indicador == "Quantidade de vendas"
                else fmt_moeda(valor)
            )

            valores = [
                indicador,
                valor_formatado,
                fmt_data(inicio),
                fmt_data(fim),
            ]

            for coluna, item in enumerate(valores):
                self.tabela_financeiro.setItem(
                    row,
                    coluna,
                    QTableWidgetItem(str(item)),
                )

    def _atualizar_producao(self, inicio, fim):
        dados = self.service.producao(inicio, fim)

        self.tabela_producao.setRowCount(0)

        for ordem in dados:
            row = self.tabela_producao.rowCount()
            self.tabela_producao.insertRow(row)

            valores = [
                ordem.get("numero", ""),
                ordem.get("produto_nome", ""),
                ordem.get("quantidade_planejada", 0),
                ordem.get("quantidade_produzida", 0),
                ordem.get("perda", 0),
                fmt_data(ordem.get("data_programada")),
                ordem.get("status", ""),
            ]

            for coluna, valor in enumerate(valores):
                self.tabela_producao.setItem(
                    row,
                    coluna,
                    QTableWidgetItem(str(valor)),
                )

    def _atualizar_clientes(self, inicio, fim):
        dados = self.service.ranking_clientes(inicio, fim)

        self.tabela_clientes.setRowCount(0)

        for item in dados:
            row = self.tabela_clientes.rowCount()
            self.tabela_clientes.insertRow(row)

            quantidade = int(item.get("quantidade_vendas", 0))
            total = float(item.get("total", 0))
            ticket = total / quantidade if quantidade else 0

            valores = [
                item.get("cliente_nome", ""),
                quantidade,
                fmt_moeda(total),
                fmt_moeda(ticket),
            ]

            for coluna, valor in enumerate(valores):
                self.tabela_clientes.setItem(
                    row,
                    coluna,
                    QTableWidgetItem(str(valor)),
                )

    def _dados_tabela_atual(self):
        tabela = self.tabs.currentWidget()
        titulo = self.tabs.tabText(
            self.tabs.currentIndex()
        )

        cabecalhos = [
            tabela.horizontalHeaderItem(coluna).text()
            for coluna in range(tabela.columnCount())
        ]

        linhas = []

        for row in range(tabela.rowCount()):
            linha = []

            for coluna in range(tabela.columnCount()):
                item = tabela.item(row, coluna)
                linha.append(item.text() if item else "")

            linhas.append(linha)

        return titulo, cabecalhos, linhas

    def _exportar_excel(self):
        titulo, cabecalhos, linhas = (
            self._dados_tabela_atual()
        )

        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar relatório Excel",
            f"{titulo.lower().replace(' ', '_')}.xlsx",
            "Excel (*.xlsx)",
        )

        if not caminho:
            return

        if not caminho.lower().endswith(".xlsx"):
            caminho += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]

        preenchimento = PatternFill(
            start_color="E8590C",
            end_color="E8590C",
            fill_type="solid",
        )
        fonte = Font(color="FFFFFF", bold=True)

        ws.append(cabecalhos)

        for celula in ws[1]:
            celula.fill = preenchimento
            celula.font = fonte

        for linha in linhas:
            ws.append(linha)

        for coluna in ws.columns:
            maior = max(
                len(str(celula.value or ""))
                for celula in coluna
            )
            ws.column_dimensions[
                coluna[0].column_letter
            ].width = min(maior + 3, 45)

        wb.save(caminho)

        info(
            self,
            "Relatório exportado",
            f"Arquivo salvo em:\n{caminho}",
        )

    def _exportar_pdf(self):
        titulo, cabecalhos, linhas = (
            self._dados_tabela_atual()
        )

        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar relatório PDF",
            f"{titulo.lower().replace(' ', '_')}.pdf",
            "PDF (*.pdf)",
        )

        if not caminho:
            return

        if not caminho.lower().endswith(".pdf"):
            caminho += ".pdf"

        try:
            documento = SimpleDocTemplate(
                caminho,
                pagesize=landscape(A4),
                rightMargin=1 * cm,
                leftMargin=1 * cm,
                topMargin=1 * cm,
                bottomMargin=1 * cm,
            )

            estilos = getSampleStyleSheet()
            elementos = [
                Paragraph(
                    EMPRESA.get(
                        "nome_fantasia",
                        "Seu Caldo 24",
                    ),
                    estilos["Title"],
                ),
                Paragraph(
                    titulo,
                    estilos["Heading2"],
                ),
                Spacer(1, 0.4 * cm),
            ]

            dados = [cabecalhos] + linhas

            tabela = Table(
                dados,
                repeatRows=1,
            )

            tabela.setStyle(TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#E8590C"),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#F1F1F1"),
                    ],
                ),
            ]))

            elementos.append(tabela)
            documento.build(elementos)

            info(
                self,
                "PDF gerado",
                f"Arquivo salvo em:\n{caminho}",
            )
        except Exception as exc:
            alerta(
                self,
                "Erro ao gerar PDF",
                str(exc),
            )