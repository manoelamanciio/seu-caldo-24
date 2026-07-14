# -*- coding: utf-8 -*-
import datetime
from PySide6.QtWidgets import QMessageBox, QFileDialog
import openpyxl
from openpyxl.styles import Font, PatternFill


def fmt_moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return "R$ 0,00"


def fmt_data(dt):
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%d/%m/%Y %H:%M")
    if isinstance(dt, datetime.date):
        return dt.strftime("%d/%m/%Y")
    return str(dt or "")


def alerta(parent, titulo, texto):
    QMessageBox.warning(parent, titulo, texto)


def info(parent, titulo, texto):
    QMessageBox.information(parent, titulo, texto)


def confirmar(parent, titulo, texto):
    resp = QMessageBox.question(parent, titulo, texto, QMessageBox.Yes | QMessageBox.No)
    return resp == QMessageBox.Yes


def exportar_para_excel(parent, cabecalhos, linhas, nome_sugerido="relatorio.xlsx", titulo_planilha="Relatório"):
    """Exporta uma tabela simples (lista de listas) para um arquivo .xlsx escolhido pelo usuário."""
    caminho, _ = QFileDialog.getSaveFileName(parent, "Salvar relatório", nome_sugerido, "Excel (*.xlsx)")
    if not caminho:
        return None
    if not caminho.endswith(".xlsx"):
        caminho += ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_planilha[:31] if titulo_planilha else "Relatório"

    header_fill = PatternFill(start_color="E8590C", end_color="E8590C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(cabecalhos)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for linha in linhas:
        ws.append(linha)

    for coluna in ws.columns:
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=10)
        ws.column_dimensions[coluna[0].column_letter].width = min(largura + 4, 45)

    wb.save(caminho)
    info(parent, "Exportado", f"Relatório exportado com sucesso para:\n{caminho}")
    return caminho